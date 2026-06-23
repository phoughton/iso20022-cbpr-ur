"""Authoring aid (NOT shipped in the package).

Reads a CBPR+ usage-guideline .xlsx and provides:
  * the Rules sheet (Index / Name / Description / Formal Rule Definition)
  * a map from the formal-rule long-name paths to the short XML-tag paths,
    reconstructed from the Full_View sheet.

Used while hand-authoring the Python rule modules so the short XML paths quoted
in rules match the published guideline exactly.
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional

import openpyxl


@dataclass
class RuleRow:
    index: str
    name: str
    description: str
    formal: str

    @property
    def is_formal(self) -> bool:
        return "FormalRule" in self.name

    @property
    def is_textual(self) -> bool:
        return "TextualRule" in self.name


def _norm(name: str) -> str:
    """Normalise a Full_View display name to the formal-rule token form."""
    name = re.sub(r"\(.*?\)", "", name)  # drop "(head.001.001.02)" etc.
    return re.sub(r"\s+", "", name)


def _sheet_rows(ws) -> List[List[str]]:
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if c is None else str(c) for c in row])
    return out


def _find_sheet(wb, *candidates):
    lower = {s.lower(): s for s in wb.sheetnames}
    for c in candidates:
        if c.lower() in lower:
            return wb[lower[c.lower()]]
    # fuzzy: contains
    for name in wb.sheetnames:
        for c in candidates:
            if c.lower() in name.lower():
                return wb[name]
    return None


class Guideline:
    def __init__(self, path: str):
        self.path = path
        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self.rules = self._load_rules()
        self.long_to_xpath = self._build_path_map()
        self.base_message = self._base_message()
        self.doc_root = self._doc_root()

    def _base_message(self):
        """e.g. 'pacs.008.001.08' - used to build the Document namespace."""
        ws = _find_sheet(self.wb, "Rules", "General Information")
        if ws is None:
            return None
        for row in _sheet_rows(ws):
            for i, cell in enumerate(row):
                if cell.strip() == "Base Message" and i + 1 < len(row):
                    return row[i + 1].strip()
        return None

    @property
    def namespace(self):
        return f"urn:iso:std:iso:20022:tech:xsd:{self.base_message}" if self.base_message else None

    def _doc_root(self):
        """Short tag of the message root under <Document>, e.g. 'FIToFICstmrCdtTrf'."""
        for v in self.long_to_xpath.values():
            if v.startswith("/Document/"):
                parts = v.split("/")
                if len(parts) > 2:
                    return parts[2]
        return None

    def _load_rules(self) -> List[RuleRow]:
        ws = _find_sheet(self.wb, "Rules")
        if ws is None:
            return []
        rows = _sheet_rows(ws)
        out = []
        for r in rows:
            idx = (r[0] if len(r) > 0 else "").strip()
            if not re.match(r"R\d+$", idx):
                continue
            out.append(
                RuleRow(
                    index=idx,
                    name=(r[1] if len(r) > 1 else "").strip(),
                    description=(r[2] if len(r) > 2 else "").strip(),
                    formal=(r[3] if len(r) > 3 else "").strip(),
                )
            )
        return out

    def _build_path_map(self) -> Dict[str, str]:
        ws = _find_sheet(self.wb, "Full_View", "FullView", "Full View")
        if ws is None:
            return {}
        rows = _sheet_rows(ws)
        # locate columns from the header row
        header = rows[0]
        col = {h.strip(): i for i, h in enumerate(header)}
        c_lvl = col.get("Lvl", 1)
        c_name = col.get("Name", 2)
        c_tag = col.get("XML Tag", 3)
        c_path = col.get("XML Path", 20)

        long_stack: Dict[int, str] = {}
        short_stack: Dict[int, str] = {}
        mapping: Dict[str, str] = {}

        for r in rows[1:]:
            name = (r[c_name] if len(r) > c_name else "").strip()
            tag = (r[c_tag] if len(r) > c_tag else "").strip().strip("<>")
            path = (r[c_path] if len(r) > c_path else "").strip()
            lvl_raw = (r[c_lvl] if len(r) > c_lvl else "").strip()
            if not name:
                continue
            # element rows have a tag or an absolute path, or are a structural
            # anchor ("Full Message" / "Document") that carries neither.
            is_element = bool(tag) or bool(path) or name in ("Full Message", "Document")
            if not is_element:
                continue
            if not tag and name == "Document":
                tag = "Document"
            try:
                lvl = int(float(lvl_raw))
            except ValueError:
                continue
            long_stack[lvl] = _norm(name)
            short_stack[lvl] = tag
            for k in [k for k in long_stack if k > lvl]:
                del long_stack[k]
                short_stack.pop(k, None)

            if path:
                short = path
            else:
                parts = [short_stack[k] for k in sorted(short_stack) if short_stack[k]]
                short = "/" + "/".join(parts) if parts else ""
            long = "/".join(long_stack[k] for k in sorted(long_stack))
            if short:
                mapping[long] = short
                # also key without the leading "Full Message" anchor
                if long.startswith("Full Message/"):
                    mapping[long[len("Full Message/"):]] = short
                # and keyed from the message root (for relative "for each" paths)
                if long.startswith("Full Message/Document/"):
                    mapping[long[len("Full Message/Document/"):]] = short
        return mapping

    def translate(self, long_path: str) -> Optional[str]:
        key = _norm(long_path.strip().strip("[]"))
        return self.long_to_xpath.get(key) or self.long_to_xpath.get(
            "Full Message/" + key if not key.startswith("Full Message/") else key
        )


def _demo(path: str):
    g = Guideline(path)
    print(f"{len(g.rules)} rules; {len(g.long_to_xpath)} path mappings")
    sample = [
        "Full Message/BusinessApplicationHeaderV02/From/FinancialInstitutionIdentification/FinancialInstitutionIdentification/BICFI",
        "Full Message/Document/FIToFICustomerCreditTransferV08/CreditTransferTransactionInformation/PaymentTypeInformation/InstructionPriority",
        "Full Message/BusinessApplicationHeaderV02/Priority",
        "Full Message/BusinessApplicationHeaderV02/CopyDuplicate",
    ]
    for s in sample:
        print(f"  {s}\n    -> {g.translate(s)}")


if __name__ == "__main__":
    _demo(sys.argv[1])
